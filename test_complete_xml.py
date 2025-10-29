#!/usr/bin/env python3
"""
Test end-to-end XML functionality by processing a sample file and exporting to XML
"""

import sys
import os

# Add the src directory to the Python path
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))

from src.excel_processor import ExcelProcessor

def test_end_to_end_xml():
    """Test complete workflow: Excel processing -> XML export."""
    
    processor = ExcelProcessor()
    
    print("🧪 Testing End-to-End XML Export...")
    print("=" * 50)
    
    # First, create a sample attendance file
    print("📝 Creating sample attendance file...")
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # Header rows
    ws['A1'] = "Date"
    ws['B1'] = "09-10-2025"
    
    ws['A2'] = "Company Name"  
    ws['B2'] = "LIGHT"
    
    ws['A3'] = "Narration"
    ws['B3'] = "Test attendance for XML export"
    
    # Row 5: Employee headers
    ws['A5'] = "EMPL NO"
    ws['B5'] = "EMPLOYEE NAME"
    ws['C5'] = "Attendance/Production Types"
    ws['D5'] = "Attendance Days"
    
    # Sample employee data
    employees = [
        [1, "Ritesh", "Present", 23],
        [2, "Milan", "Absent", 2], 
        [3, "Anil", "Overtime @ 1.25", 23],
        [4, "Utkarsh", "Overtime @ 1.50", 12]
    ]
    
    for i, emp in enumerate(employees, start=6):
        ws[f'A{i}'] = emp[0]
        ws[f'B{i}'] = emp[1]
        ws[f'C{i}'] = emp[2]
        ws[f'D{i}'] = emp[3]
    
    # Save file
    test_file = "test_sample_attendance.xlsx"
    wb.save(test_file)
    print(f"✅ Sample file created: {test_file}")
    
    # Process the attendance file
    print("\n⚙️ Processing attendance file...")
    result = processor.process_attendance_sheet(test_file)
    
    if result.get("success"):
        print("✅ Attendance processing successful!")
        print(f"   📊 Found {result.get('total_employees', 0)} employees")
        print(f"   🏢 Company: {result.get('company_name', 'N/A')}")
        print(f"   📅 Date: {result.get('date', 'N/A')}")
        
        # Export to XML
        print("\n📤 Exporting to XML...")
        xml_file = "test_complete_export.xml"
        xml_result = processor.export_attendance_xml(result, xml_file)
        
        if xml_result:
            print(f"✅ XML export successful: {xml_result}")
            
            # Read and validate XML
            with open(xml_result, 'r', encoding='utf-8') as f:
                xml_content = f.read()
            
            print("\n🔍 XML Validation:")
            print(f"   ✅ File size: {len(xml_content)} characters")
            print(f"   ✅ Contains company: {'LIGHT' in xml_content}")
            print(f"   ✅ Contains employees: {xml_content.count('<ATTENDANCEENTRIES.LIST>') == 4}")
            print(f"   ✅ Valid XML structure: {'<ENVELOPE>' in xml_content and '</ENVELOPE>' in xml_content}")
            
        else:
            print("❌ XML export failed")
    
    else:
        print("❌ Attendance processing failed:")
        print(f"   Error: {result.get('error', 'Unknown error')}")
    
    # Cleanup
    try:
        os.remove(test_file)
        print(f"\n🧹 Cleaned up test file: {test_file}")
    except:
        pass
    
    print("\n" + "=" * 50)
    print("🏁 End-to-End XML Test Complete")

if __name__ == "__main__":
    test_end_to_end_xml()