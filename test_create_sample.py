#!/usr/bin/env python3
"""
Create a sample attendance Excel file for testing
This creates an Excel file matching the structure shown in the user's image
"""

import pandas as pd
import os

def create_sample_attendance_excel():
    """Create a sample attendance Excel file matching the expected format."""
    
    # Use openpyxl to create Excel file with exact structure
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # Row 1: Date in column B
    ws['A1'] = "Date"
    ws['B1'] = "09-10-2025"
    
    # Row 2: Company Name in column B
    ws['A2'] = "Company Name"
    ws['B2'] = "LIGHT"
    
    # Row 3: Narration in column B
    ws['A3'] = "Narration"
    ws['B3'] = "Test attendance"
    
    # Row 4: Empty
    
    # Row 5: Employee headers
    ws['A5'] = "EMPL NO"
    ws['B5'] = "EMPLOYEE NAME"
    ws['C5'] = "Attendance/Production Types"
    ws['D5'] = "Attendance Days"
    
    # Employee data starting from row 6
    employees = [
        [1, "Ritesh", "Present", 23],
        [2, "Milan", "Absent", 22], 
        [3, "Anil", "Overtime @ 1.25", 23],
        [4, "Utkarsh", "Overtime @ 1.50", 12],
        [5, "John", "Present", 25],
        [6, "Sarah", "Half Day", 15],
        [7, "Mike", "Overtime @ 1.25", 28],
        [8, "Lisa", "Present", 22]
    ]
    
    for i, emp in enumerate(employees, start=6):
        ws[f'A{i}'] = emp[0]  # Employee No
        ws[f'B{i}'] = emp[1]  # Employee Name
        ws[f'C{i}'] = emp[2]  # Attendance Type
        ws[f'D{i}'] = emp[3]  # Attendance Days
    
    # Save as Excel file
    output_file = "sample_attendance.xlsx"
    output_path = os.path.join(os.path.dirname(__file__), output_file)
    
    wb.save(output_path)
    
    print("📊 Created Excel with structure:")
    print("  Row 1: Date = 09-10-2025 (in column B)")
    print("  Row 2: Company Name = LIGHT (in column B)")
    print("  Row 3: Narration = Test attendance (in column B)")
    print("  Row 4: Empty")
    print("  Row 5: Employee headers")
    print("  Row 6+: Employee data")
    
    # Verify what was saved
    test_df = pd.read_excel(output_path)
    print(f"\n📋 Verification - saved Excel has {test_df.shape[0]} rows × {test_df.shape[1]} columns")
    print("First 3 rows:")
    for i in range(min(3, len(test_df))):
        print(f"  Row {i}: {list(test_df.iloc[i])}")
    
    print(f"✅ Sample attendance Excel file created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_sample_attendance_excel()