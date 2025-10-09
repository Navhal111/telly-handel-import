#!/usr/bin/env python3
"""
Create a sample payroll Excel file for testing
This creates an Excel file matching the structure shown in the user's image
"""

from openpyxl import Workbook
import os

def create_sample_payroll_excel():
    """Create a sample payroll Excel file matching the expected format."""
    
    wb = Workbook()
    
    ws = wb.active
    
    # Header rows
    ws['A1'] = "Date"
    ws['B1'] = "09-10-2025"
    
    ws['A2'] = "Company Name"
    ws['B2'] = "LIGHT"
    
    ws['A3'] = "Account"
    ws['B3'] = "PEAPULE BANK OF UNITE"
    
    ws['A4'] = "Narration"
    ws['B4'] = "Test attendance"
    
    # Row 6: Payroll headers
    headers = ['EMPL NO', 'EMPLOYEE NAME', 'BASIC', 'HRA', 'MEDICAL', 'RESPONSIBILITY', 'FUEL', 'ZSSF @ 7%', 'ZHSF @ 3.5%', 'PAYE']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=6, column=col, value=header)
    
    # Sample employee payroll data
    employees = [
        [1, "Ritesh", 2200000, 440000, 660000, 1000000, 250000, 154000, 154000, 823700],
        [2, "Milan", 4200000, 140000, 98000, 1000000, 0, 67000, 98250, 563100],
        [3, "John", 1800000, 360000, 540000, 800000, 200000, 126000, 126000, 690000],
        [4, "Sarah", 3500000, 120000, 85000, 900000, 0, 58000, 85500, 485000],
        [5, "Mike", 2800000, 560000, 420000, 1200000, 300000, 196000, 196000, 1050000]
    ]
    
    # Add employee data starting from row 7
    for row_idx, emp in enumerate(employees, start=7):
        for col_idx, value in enumerate(emp, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save as Excel file
    output_file = "sample_payroll.xlsx"
    output_path = os.path.join(os.path.dirname(__file__), output_file)
    
    wb.save(output_path)
    
    print("📊 Created Payroll Excel with structure:")
    print("  Row 1: Date = 09-10-2025")
    print("  Row 2: Company Name = LIGHT")
    print("  Row 3: Account = PEAPULE BANK OF UNITE")
    print("  Row 4: Narration = Test attendance")
    print("  Row 6: Payroll headers")
    print("  Row 7+: Employee payroll data")
    print(f"✅ Sample payroll Excel file created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_sample_payroll_excel()