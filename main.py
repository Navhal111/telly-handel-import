import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import sys
import os
import threading
import subprocess
import platform

# Add the src directory to the Python path
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))

from src.excel_processor import ExcelProcessor


class ModernExcelProcessor:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel File Processor - Attendance & Payroll")
        self.root.geometry("1000x680")
        self.root.minsize(900, 600)
        
        # Configure style
        self.setup_style()
        
        # Initialize Excel processor
        self.processor = ExcelProcessor()
        
        # Create GUI
        self.create_gui()
        
        # Center the window
        self.center_window()
    
    def setup_style(self):
        """Set up modern styling for the application."""
        style = ttk.Style()
        
        # Configure colors
        self.colors = {
            'primary': '#2c3e50',
            'secondary': '#3498db',
            'success': '#27ae60',
            'warning': '#f39c12',
            'danger': '#e74c3c',
            'light': '#ecf0f1',
            'dark': '#34495e',
            'white': '#ffffff'
        }
        
        # Configure root window
        self.root.configure(bg=self.colors['light'])
        
        # Configure ttk styles
        style.theme_use('clam')
        
        # Enhanced button styles with optimized padding
        style.configure('Primary.TButton',
                       background=self.colors['primary'],
                       foreground=self.colors['white'],
                       font=('Arial', 11, 'bold'),
                       padding=(20, 12),
                       borderwidth=0)
        
        style.configure('Secondary.TButton',
                       background=self.colors['secondary'],
                       foreground=self.colors['white'],
                       font=('Arial', 11, 'bold'),
                       padding=(20, 12),
                       borderwidth=0)
        
        style.configure('Success.TButton',
                       background=self.colors['success'],
                       foreground=self.colors['white'],
                       font=('Arial', 11, 'bold'),
                       padding=(20, 12),
                       borderwidth=0)
        
        # Button hover effects
        style.map('Primary.TButton',
                 background=[('active', '#1a252f')])
        style.map('Secondary.TButton', 
                 background=[('active', '#2980b9')])
        style.map('Success.TButton',
                 background=[('active', '#229954')])
        
        # Configure label styles with bigger title
        style.configure('Title.TLabel',
                       background=self.colors['light'],
                       foreground=self.colors['primary'],
                       font=('Arial', 20, 'bold'))
        
        style.configure('Subtitle.TLabel',
                       background=self.colors['light'],
                       foreground=self.colors['dark'],
                       font=('Arial', 12))
        
        style.configure('Info.TLabel',
                       background=self.colors['light'],
                       foreground=self.colors['dark'],
                       font=('Arial', 10))
    
    def center_window(self):
        """Center the window on the screen."""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        pos_x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        pos_y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{pos_x}+{pos_y}")
    
    def create_gui(self):
        """Create the main GUI interface."""
        # Main container with optimized spacing
        self.main_frame = tk.Frame(self.root, bg=self.colors['light'], padx=20, pady=15)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame = self.main_frame  # Keep local reference for existing code
        
        # Header section with reduced spacing
        header_frame = tk.Frame(main_frame, bg=self.colors['light'])
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Title with icon - bigger and more prominent
        title_label = ttk.Label(header_frame, text="� Excel File Processor", style='Title.TLabel')
        title_label.pack(pady=(0, 5))
        
        subtitle_label = ttk.Label(header_frame, text="Upload and process Attendance or Payroll Excel files with intelligent data extraction", style='Subtitle.TLabel')
        subtitle_label.pack(pady=(0, 3))
        
        # Description line
        desc_label = ttk.Label(header_frame, text="Supports header extraction, employee data processing, and JSON export", style='Info.TLabel')
        desc_label.pack()
        
        # Upload sections container with reduced spacing
        upload_frame = tk.Frame(main_frame, bg=self.colors['light'])
        upload_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Attendance upload section
        self.create_upload_section(upload_frame, "Attendance Sheet", "attendance", 0)
        
        # Professional separator with text - reduced spacing
        separator_frame = tk.Frame(upload_frame, bg=self.colors['light'])
        separator_frame.pack(fill=tk.X, pady=15)
        
        sep_line1 = tk.Frame(separator_frame, height=1, bg=self.colors['dark'])
        sep_line1.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        sep_label = ttk.Label(separator_frame, text=" OR ", style='Info.TLabel')
        sep_label.pack(side=tk.LEFT, padx=10)
        
        sep_line2 = tk.Frame(separator_frame, height=1, bg=self.colors['dark'])
        sep_line2.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        
        # Payroll upload section
        self.create_upload_section(upload_frame, "Payroll Sheet", "payroll", 1)
        
        # Status area with optimized spacing
        status_container = tk.Frame(main_frame, bg=self.colors['light'])
        status_container.pack(fill=tk.BOTH, expand=True, pady=(15, 0))
        
        status_label = ttk.Label(status_container, text="📊 Status:", style='Subtitle.TLabel')
        status_label.pack(anchor=tk.W, pady=(0, 8))
        
        # Status display area with optimized styling
        self.status_frame = tk.Frame(status_container, bg=self.colors['white'], relief=tk.SOLID, bd=1)
        self.status_frame.pack(fill=tk.BOTH, expand=True, padx=5, ipady=10)
        
        # Status icon and text container
        status_content = tk.Frame(self.status_frame, bg=self.colors['white'])
        status_content.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        self.status_text = tk.Label(
            status_content,
            text="✨ Ready to process Excel files...",
            font=('Arial', 12),
            bg=self.colors['white'],
            fg=self.colors['dark'],
            wraplength=700,
            justify=tk.LEFT
        )
        self.status_text.pack(anchor=tk.W)
    
    def create_upload_section(self, parent, title, section_type, row):
        """Create an upload section for either Attendance or Payroll."""
        # Main section container with optimized spacing
        section_container = tk.Frame(parent, bg=self.colors['light'])
        section_container.pack(fill=tk.X, pady=10, padx=15)
        
        section_frame = tk.Frame(section_container, bg=self.colors['white'], relief=tk.SOLID, bd=1)
        section_frame.pack(fill=tk.X, ipady=15, ipadx=20)
        
        # Section title with icon - reduced spacing
        title_frame = tk.Frame(section_frame, bg=self.colors['white'])
        title_frame.pack(fill=tk.X, pady=(0, 10))
        
        icon = "📊" if section_type == "attendance" else "💰"
        title_text = f"{icon} {title}"
        title_label = ttk.Label(title_frame, text=title_text, style='Subtitle.TLabel')
        title_label.configure(background=self.colors['white'])
        title_label.pack(anchor=tk.W)
        
        # File path display with optimized spacing
        file_frame = tk.Frame(section_frame, bg=self.colors['white'])
        file_frame.pack(fill=tk.X, pady=(0, 12))
        
        path_label = ttk.Label(file_frame, text="📁 Selected file:", style='Info.TLabel')
        path_label.configure(background=self.colors['white'])
        path_label.pack(anchor=tk.W, pady=(0, 5))
        
        # Store file path variable
        file_path_var = tk.StringVar(value="No file selected")
        setattr(self, f"{section_type}_file_path", file_path_var)
        
        # File path display with background
        path_container = tk.Frame(file_frame, bg=self.colors['light'], relief=tk.SUNKEN, bd=1)
        path_container.pack(fill=tk.X, pady=(0, 5))
        
        path_display = ttk.Label(
            path_container,
            textvariable=file_path_var,
            style='Info.TLabel',
            width=70
        )
        path_display.configure(background=self.colors['light'])
        path_display.pack(anchor=tk.W, padx=10, pady=8)
        
        # Buttons frame
        btn_frame = tk.Frame(section_frame, bg=self.colors['white'])
        btn_frame.pack(fill=tk.X)
        
        # Browse button with icon
        browse_btn = ttk.Button(
            btn_frame,
            text="📂 Browse File",
            command=lambda: self.browse_file(section_type),
            style='Primary.TButton'
        )
        browse_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # Process button with icon
        process_icon = "⚙️" if section_type == "attendance" else "💵"
        process_btn = ttk.Button(
            btn_frame,
            text=f"{process_icon} Process {title}",
            command=lambda: self.process_file(section_type),
            style='Success.TButton'
        )
        process_btn.pack(side=tk.LEFT, padx=(0, 20))
        
        # Visual separator
        separator = tk.Frame(btn_frame, width=2, height=30, bg=self.colors['dark'])
        separator.pack(side=tk.LEFT, padx=(0, 20))
        
        # Download Sample button with icon
        sample_btn = ttk.Button(
            btn_frame,
            text="📄 Download Sample",
            command=lambda: self.download_sample(section_type),
            style='Secondary.TButton'
        )
        sample_btn.pack(side=tk.LEFT)
    
    def browse_file(self, section_type):
        """Open file dialog to select Excel file."""
        file_types = [
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
        
        filename = filedialog.askopenfilename(
            title=f"Select {section_type.title()} Excel File",
            filetypes=file_types,
            initialdir=os.path.expanduser("~")
        )
        
        if filename:
            file_path_var = getattr(self, f"{section_type}_file_path")
            file_path_var.set(filename)
            self.add_result(f"✓ {section_type.title()} file selected: {os.path.basename(filename)}")
    
    def process_file(self, section_type):
        """Process the selected Excel file."""
        file_path_var = getattr(self, f"{section_type}_file_path")
        file_path = file_path_var.get()
        
        if file_path == "No file selected" or not file_path:
            messagebox.showwarning("No File Selected", f"Please select a {section_type} Excel file first.")
            return
        
        if not os.path.exists(file_path):
            messagebox.showerror("File Not Found", "The selected file no longer exists.")
            return
        
        # Get the process button and show loading
        process_btn = None
        for widget in self.root.winfo_children():
            for child in self._get_all_widgets(widget):
                if isinstance(child, ttk.Button) and f"Process {section_type.title()}" in child.cget('text'):
                    process_btn = child
                    break
        
        if process_btn:
            process_btn.configure(text="Processing...", state='disabled')
            self.root.update()
        
        # Update status
        self.status_text.configure(text=f"Processing {section_type} file: {os.path.basename(file_path)}...")
        self.root.update()
        
        # Process in background thread
        thread = threading.Thread(target=self._process_file_background, args=(section_type, file_path, process_btn))
        thread.daemon = True
        thread.start()
    
    def _get_all_widgets(self, widget):
        """Recursively get all widgets from a parent widget."""
        widgets = [widget]
        for child in widget.winfo_children():
            widgets.extend(self._get_all_widgets(child))
        return widgets
    
    def _process_file_background(self, section_type, file_path, process_btn):
        """Process file in background thread."""
        try:
            if section_type == "attendance":
                result = self.processor.process_attendance_sheet(file_path)
            else:
                result = self.processor.process_payroll_sheet(file_path)
            
            # Update UI in main thread
            self.root.after(0, self._handle_processing_result, section_type, result, process_btn)
            
        except Exception as e:
            error_result = {"error": str(e), "success": False}
            self.root.after(0, self._handle_processing_result, section_type, error_result, process_btn)
    
    def _handle_processing_result(self, section_type, result, process_btn):
        """Handle the processing result - show data screen or error popup."""
        print(f"🎯 DEBUG: Handling processing result for {section_type}")
        print(f"🎯 DEBUG: Result success: {result.get('success', False)}")
        print(f"🎯 DEBUG: Result keys: {list(result.keys()) if result else 'None'}")
        
        # Restore button state
        if process_btn:
            button_text = f"Process {section_type.title()} Sheet"
            process_btn.configure(text=button_text, state='normal')
        
        if result.get("success", False):
            # Valid Excel - show data screen
            print(f"✅ DEBUG: Valid Excel, showing data screen")
            self.status_text.configure(text=f"{section_type.title()} processing completed successfully!")
            self.show_data_screen(section_type, result)
        else:
            # Invalid Excel - show error popup
            print(f"❌ DEBUG: Invalid Excel, showing error popup")
            self.status_text.configure(text="Ready to process Excel files...")
            error_msg = result.get('error', 'Unknown error occurred')
            
            # Create detailed error message
            if 'validation_errors' in result:
                validation_errors = result['validation_errors']
                detailed_msg = "Invalid Excel format detected!\n\n"
                detailed_msg += "Errors found:\n"
                for error in validation_errors:
                    detailed_msg += f"• {error}\n"
                detailed_msg += "\n" + self._get_format_help_text()
            else:
                detailed_msg = f"Error processing file:\n{error_msg}\n\nPlease check your Excel file and try again."
            
            messagebox.showerror("Invalid Excel File", detailed_msg)
    
    def add_result(self, text):
        """Add text to the results area."""
        self.results_text.insert(tk.END, f"{text}\n")
        self.results_text.see(tk.END)
        # Force UI update
        self.root.update_idletasks()
    
    def show_loading(self, section_type, file_name):
        """Show loading indicator in results."""
        self.results_text.insert(tk.END, f"🔄 Processing {section_type} file: {file_name}\n")
        self.results_text.insert(tk.END, "⏳ Loading... Please wait while we analyze your Excel file.\n")
        self.results_text.insert(tk.END, "   • Reading file structure...\n")
        self.results_text.insert(tk.END, "   • Analyzing data types...\n")
        self.results_text.insert(tk.END, "   • Processing columns and rows...\n\n")
        self.results_text.see(tk.END)
        self.root.update_idletasks()
        self.root.update()
    
    def clear_loading(self):
        """Clear loading messages."""
        # This will be handled by the results display
        pass
    
    def display_results(self, section_type, result):
        """Display the processing results directly."""
        print(f"🎯 display_results called with section_type: {section_type}")
        print(f"📊 Result success: {result.get('success', 'No success key')}")
        
        # Clear and start fresh
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, f"✅ PROCESSING COMPLETED FOR {section_type.upper()}!\n\n")
        
        if result.get("success", False):
            # File information
            self.results_text.insert(tk.END, f"📊 FILE INFORMATION:\n")
            self.results_text.insert(tk.END, f"   📁 File: {result['file_name']}\n")
            self.results_text.insert(tk.END, f"   📋 Type: {result['sheet_type']}\n")
            self.results_text.insert(tk.END, f"   📏 Total Rows: {result['total_rows']}\n")
            self.results_text.insert(tk.END, f"   📐 Total Columns: {result['total_columns']}\n\n")
            
            if section_type == "attendance":
                # Header Information (Date, Company, Narration)
                self.results_text.insert(tk.END, f"📅 ATTENDANCE SHEET HEADER:\n")
                self.results_text.insert(tk.END, f"   📅 Date: {result.get('date', 'Not found')}\n")
                self.results_text.insert(tk.END, f"   🏢 Company Name: {result.get('company_name', 'Not found')}\n")
                self.results_text.insert(tk.END, f"   📝 Narration: {result.get('narration', 'Not found')}\n\n")
                
                # Employee Data
                employee_data = result.get('employee_data', [])
                self.results_text.insert(tk.END, f"👥 EMPLOYEE DATA ({result.get('total_employees', 0)} employees):\n")
                
                if employee_data:
                    # Show first 10 employees
                    for i, emp in enumerate(employee_data[:10], 1):
                        self.results_text.insert(tk.END, f"   {i}. ")
                        self.results_text.insert(tk.END, f"ID: {emp.get('employee_no', 'N/A')}, ")
                        self.results_text.insert(tk.END, f"Name: {emp.get('employee_name', 'N/A')}, ")
                        self.results_text.insert(tk.END, f"Type: {emp.get('attendance_type', 'N/A')}, ")
                        self.results_text.insert(tk.END, f"Days: {emp.get('attendance_days', 'N/A')}\n")
                    
                    if len(employee_data) > 10:
                        self.results_text.insert(tk.END, f"   ... and {len(employee_data) - 10} more employees\n")
                else:
                    self.results_text.insert(tk.END, "   No employee data found\n")
                
                self.results_text.insert(tk.END, f"\n   📍 Employee data starts at row: {result.get('employee_data_start_row', 'Unknown')}\n")
                
                # Validation warnings
                warnings = result.get('validation_warnings', [])
                if warnings:
                    self.results_text.insert(tk.END, f"\n⚠️ WARNINGS:\n")
                    for warning in warnings:
                        self.results_text.insert(tk.END, f"   • {warning}\n")
            
            else:  # payroll
                # Basic file info for payroll
                columns = result.get('columns', [])
                self.results_text.insert(tk.END, f"📋 COLUMNS ({len(columns)}):\n")
                for i, col in enumerate(columns[:10], 1):
                    self.results_text.insert(tk.END, f"   {i}. {col}\n")
                if len(columns) > 10:
                    self.results_text.insert(tk.END, f"   ... and {len(columns) - 10} more columns\n")
                
                # Sample data
                sample_data = result.get('sample_data', [])
                if sample_data:
                    self.results_text.insert(tk.END, f"\n📄 SAMPLE DATA (First Row):\n")
                    first_row = sample_data[0]
                    for key, value in list(first_row.items())[:5]:
                        self.results_text.insert(tk.END, f"   {key}: {value}\n")
        
        else:
            # Error handling
            self.results_text.insert(tk.END, f"❌ ERROR PROCESSING FILE:\n")
            self.results_text.insert(tk.END, f"   {result.get('error', 'Unknown error')}\n\n")
            
            # Show validation errors if available
            validation_errors = result.get('validation_errors', [])
            if validation_errors:
                self.results_text.insert(tk.END, f"� VALIDATION ERRORS:\n")
                for error in validation_errors:
                    self.results_text.insert(tk.END, f"   • {error}\n")
                
                self.results_text.insert(tk.END, f"\n💡 EXPECTED FORMAT FOR ATTENDANCE SHEET:\n")
                self.results_text.insert(tk.END, f"   • Row 1: Date in column B (e.g., 09-10-2025)\n")
                self.results_text.insert(tk.END, f"   • Row 2: Company Name in column B (e.g., LIGHT)\n")
                self.results_text.insert(tk.END, f"   • Row 3: Narration in column B (e.g., Test attendance)\n")
                self.results_text.insert(tk.END, f"   • Row 5+: Employee data with headers:\n")
                self.results_text.insert(tk.END, f"     - EMPL NO\n")
                self.results_text.insert(tk.END, f"     - EMPLOYEE NAME\n")
                self.results_text.insert(tk.END, f"     - Attendance/Production Types\n")
                self.results_text.insert(tk.END, f"     - Attendance Days\n")
        
        self.results_text.insert(tk.END, "\n" + "="*80 + "\n\n")
        self.results_text.see(tk.END)
        self.results_text.update()  # Force display update

    def add_result(self, text):
        """Update status text (replacing old results functionality)."""
        self.status_text.configure(text=text)
    
    def download_sample(self, section_type):
        """Download and open a sample Excel file for the specified section type."""
        try:
            sample_file = None
            
            if section_type == "attendance":
                # Create attendance sample
                sample_file = self.create_attendance_sample()
                sample_name = "Sample_Attendance.xlsx"
            elif section_type == "payroll":
                # Create payroll sample
                sample_file = self.create_payroll_sample()
                sample_name = "Sample_Payroll.xlsx"
            
            if sample_file:
                # Show success message
                messagebox.showinfo(
                    "Sample Downloaded", 
                    f"Sample {section_type} file created successfully!\n\n"
                    f"File: {sample_name}\n"
                    f"Location: {sample_file}\n\n"
                    f"The file has been opened for you to view the expected format."
                )
                
                # Update status
                self.add_result(f"📁 Sample {section_type} file downloaded: {sample_name}")
                
                # Open the file with default application
                import subprocess
                import platform
                
                if platform.system() == 'Darwin':  # macOS
                    subprocess.run(['open', sample_file])
                elif platform.system() == 'Windows':  # Windows
                    subprocess.run(['start', sample_file], shell=True)
                else:  # Linux
                    subprocess.run(['xdg-open', sample_file])
            
        except Exception as e:
            messagebox.showerror("Download Error", f"Error creating sample file: {str(e)}")
    
    def create_attendance_sample(self):
        """Create a sample attendance Excel file."""
        from openpyxl import Workbook
        import os
        
        wb = Workbook()
        ws = wb.active
        
        # Header rows
        ws['A1'] = "Date"
        ws['B1'] = "09-10-2025"
        
        ws['A2'] = "Company Name"
        ws['B2'] = "LIGHT"
        
        ws['A3'] = "Narration"
        ws['B3'] = "Test attendance"
        
        # Row 5: Employee headers
        ws['A5'] = "EMPL NO"
        ws['B5'] = "EMPLOYEE NAME"
        ws['C5'] = "Attendance/Production Types"
        ws['D5'] = "Attendance Days"
        
        # Sample employee data
        employees = [
            [1, "Ritesh", "Present", 23],
            [2, "Milan", "Absent", 22], 
            [3, "Anil", "Overtime @ 1.25", 23],
            [4, "Utkarsh", "Overtime @ 1.50", 12],
            [5, "John", "Present", 25],
            [6, "Sarah", "Half Day", 15]
        ]
        
        for i, emp in enumerate(employees, start=6):
            ws[f'A{i}'] = emp[0]
            ws[f'B{i}'] = emp[1]
            ws[f'C{i}'] = emp[2]
            ws[f'D{i}'] = emp[3]
        
        # Save file
        output_path = os.path.join(os.path.expanduser("~/Downloads"), "Sample_Attendance.xlsx")
        wb.save(output_path)
        return output_path
    
    def create_payroll_sample(self):
        """Create a sample payroll Excel file."""
        from openpyxl import Workbook
        import os
        
        wb = Workbook()
        ws = wb.active
        
        # Header rows
        ws['A1'] = "Date"
        ws['B1'] = "09-10-2025"
        
        ws['A2'] = "Company Name"
        ws['B2'] = "LIGHT"
        
        ws['A3'] = "Account"
        ws['B3'] = "PEAPULE BANK OF UNITE"
        
        ws['A4'] = "Narration"
        ws['B4'] = "Monthly Payroll"
        
        # Row 6: Payroll headers
        headers = ['EMPL NO', 'EMPLOYEE NAME', 'BASIC', 'HRA', 'MEDICAL', 'RESPONSIBILITY', 'FUEL', 'ZSSF @ 7%', 'ZHSF @ 3.5%', 'PAYE']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=6, column=col, value=header)
        
        # Sample payroll data
        employees = [
            [1, "Ritesh", 2200000, 440000, 660000, 1000000, 250000, 154000, 154000, 823700],
            [2, "Milan", 4200000, 140000, 98000, 1000000, 0, 67000, 98250, 563100],
            [3, "John", 1800000, 360000, 540000, 800000, 200000, 126000, 126000, 690000],
            [4, "Sarah", 3500000, 120000, 85000, 900000, 0, 58000, 85500, 485000],
            [5, "Mike", 2800000, 560000, 420000, 1200000, 300000, 196000, 196000, 1050000]
        ]
        
        for row_idx, emp in enumerate(employees, start=7):
            for col_idx, value in enumerate(emp, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save file
        output_path = os.path.join(os.path.expanduser("~/Downloads"), "Sample_Payroll.xlsx")
        wb.save(output_path)
        return output_path
    
    def show_main_screen(self):
        """Show the main screen (hide data screen)."""
        # Remove data frame if it exists
        if hasattr(self, 'data_frame'):
            self.data_frame.destroy()
            delattr(self, 'data_frame')
        
        # Clear all widgets from root
        for widget in self.root.winfo_children():
            widget.destroy()
        
        # Recreate main interface
        self.create_gui()
    
    def export_data_from_screen(self, section_type, result):
        """Export data from the data screen."""
        if not hasattr(self, 'current_result') or not self.current_result:
            messagebox.showerror("Export Error", "No data available to export.")
            return
            
        result = self.current_result
        section_type = self.current_section_type
        
        file_types = [("JSON files", "*.json"), ("All files", "*.*")]
        default_name = f"{os.path.splitext(result['file_name'])[0]}_processed_{section_type}.json"
        
        filename = filedialog.asksaveasfilename(
            title=f"Save {section_type.title()} Data",
            defaultextension=".json",
            filetypes=file_types,
            initialfile=default_name
        )
        
        if filename:
            try:
                if section_type == "attendance":
                    export_path = self.processor.export_attendance_data(result, filename)
                elif section_type == "payroll":
                    export_path = self.processor.export_payroll_data(result, filename)
                else:
                    import json
                    with open(filename, 'w', encoding='utf-8') as f:
                        json.dump(result, f, indent=2, ensure_ascii=False)
                    export_path = filename
                
                if export_path:
                    messagebox.showinfo("Export Successful", 
                                      f"Data exported successfully to:\n{export_path}")
                else:
                    messagebox.showerror("Export Failed", "Failed to export data.")
                    
            except Exception as e:
                messagebox.showerror("Export Error", f"Error exporting data: {str(e)}")
    
    def _get_format_help_text(self):
        """Get help text for expected format."""
        return """Expected format for Excel Sheets:

For Attendance Sheet:
• Row 1: Date in column B (e.g., 09-10-2025)
• Row 2: Company Name in column B (e.g., LIGHT)
• Row 3: Narration in column B (e.g., Test attendance)
• Row 5+: Employee data with headers:
  - EMPL NO, EMPLOYEE NAME, Attendance/Production Types, Attendance Days

For Payroll Sheet:
• Row 1: Date in column B (e.g., 09-10-2025)
• Row 2: Company Name in column B (e.g., LIGHT)
• Row 3: Account in column B (e.g., PEAPULE BANK OF UNITE)
• Row 6+: Employee data with headers:
  - EMPL NO, EMPLOYEE NAME, BASIC, HRA, MEDICAL, etc."""
    
    def show_data_screen(self, section_type, result):
        """Show data screen in the same window."""
        # Hide main content
        for widget in self.root.winfo_children():
            widget.pack_forget()
        
        # Create data screen frame with better layout
        self.data_frame = tk.Frame(self.root, bg=self.colors['light'])
        self.data_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Title
        title_text = f"{section_type.title()} Data Extracted Successfully"
        title_label = ttk.Label(self.data_frame, text=title_text, style='Title.TLabel')
        title_label.pack(pady=(0, 15))
        
        # Create main content frame (no scrollbar for now to fix layout)
        content_frame = tk.Frame(self.data_frame, bg=self.colors['light'])
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Add content directly to content frame
        self._populate_data_screen(content_frame, section_type, result)
        
        # Bottom buttons with better layout
        btn_frame = tk.Frame(self.data_frame, bg=self.colors['light'])
        btn_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(15, 5))
        
        # Left side buttons
        left_btn_frame = tk.Frame(btn_frame, bg=self.colors['light'])
        left_btn_frame.pack(side=tk.LEFT)
        
        # Back button
        back_btn = ttk.Button(
            left_btn_frame,
            text="← Back",
            command=self.show_main_screen,
            style='Secondary.TButton'
        )
        back_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Export button
        export_btn = ttk.Button(
            left_btn_frame,
            text="Export to JSON",
            command=lambda: self.export_data_from_screen(section_type, result),
            style='Success.TButton'
        )
        export_btn.pack(side=tk.LEFT)
        
        # Right side buttons
        right_btn_frame = tk.Frame(btn_frame, bg=self.colors['light'])
        right_btn_frame.pack(side=tk.RIGHT)
        
        # Cancel button
        cancel_btn = ttk.Button(
            right_btn_frame,
            text="Cancel",
            command=self.show_main_screen,
            style='Secondary.TButton'
        )
        cancel_btn.pack(side=tk.RIGHT)
        
        # Store current result for export
        self.current_result = result
        self.current_section_type = section_type
    
    def _populate_data_screen(self, parent_frame, section_type, result):
        """Populate the data screen with extracted information."""
        # Configure parent frame to be responsive
        parent_frame.columnconfigure(0, weight=1)
        
        # Debug: Print what we're displaying
        print(f"🎯 DEBUG: Populating {section_type} screen")
        print(f"🎯 DEBUG: Result keys: {list(result.keys()) if result else 'None'}")
        print(f"🎯 DEBUG: Success: {result.get('success', False)}")
        
        if section_type == "attendance":
            # File Information Section
            file_section = tk.LabelFrame(parent_frame, text="📊 File Information", 
                                       font=('Arial', 11, 'bold'), bg=self.colors['light'],
                                       fg=self.colors['primary'])
            file_section.pack(fill=tk.X, pady=(0, 10), padx=5)
            
            file_info = tk.Frame(file_section, bg=self.colors['light'])
            file_info.pack(fill=tk.X, padx=10, pady=8)
            
            tk.Label(file_info, text=f"📁 File: {result.get('file_name', 'N/A')}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['dark']).pack(anchor=tk.W, pady=2)
            tk.Label(file_info, text=f"📏 Total Rows: {result.get('total_rows', 0)}", 
                    font=('Arial', 10), bg=self.colors['light'], fg=self.colors['dark']).pack(anchor=tk.W, pady=2)
            tk.Label(file_info, text=f"📐 Total Columns: {result.get('total_columns', 0)}", 
                    font=('Arial', 10), bg=self.colors['light'], fg=self.colors['dark']).pack(anchor=tk.W, pady=2)
            
            # Header Information Section
            header_section = tk.LabelFrame(parent_frame, text="📋 Header Information", 
                                         font=('Arial', 11, 'bold'), bg=self.colors['light'],
                                         fg=self.colors['primary'])
            header_section.pack(fill=tk.X, pady=(0, 10), padx=5)
            
            header_info = tk.Frame(header_section, bg=self.colors['light'])
            header_info.pack(fill=tk.X, padx=10, pady=8)
            
            # Debug print header values
            print(f"🔍 DEBUG: Date: {result.get('date', 'Not found')}")
            print(f"🔍 DEBUG: Company: {result.get('company_name', 'Not found')}")
            print(f"🔍 DEBUG: Narration: {result.get('narration', 'Not found')}")
            
            tk.Label(header_info, text=f"📅 Date: {result.get('date', 'Not found')}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['success']).pack(anchor=tk.W, pady=3)
            tk.Label(header_info, text=f"🏢 Company Name: {result.get('company_name', 'Not found')}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['success']).pack(anchor=tk.W, pady=3)
            tk.Label(header_info, text=f"📝 Narration: {result.get('narration', 'Not found')}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['success']).pack(anchor=tk.W, pady=3)
            
            # Employee Data Section
            employee_data = result.get('employee_data', [])
            print(f"🔍 DEBUG: Employee data count: {len(employee_data)}")
            if employee_data:
                print(f"🔍 DEBUG: First employee: {employee_data[0]}")
            
            emp_section = tk.LabelFrame(parent_frame, text=f"👥 Employee Data ({len(employee_data)} employees)", 
                                      font=('Arial', 11, 'bold'), bg=self.colors['light'],
                                      fg=self.colors['primary'])
            emp_section.pack(fill=tk.BOTH, expand=True, pady=(0, 10), padx=5)
            
            # Employee data table
            if employee_data:
                # Create frame for table with better sizing
                table_frame = tk.Frame(emp_section, bg=self.colors['white'])
                table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
                
                # Create treeview for employee data with better sizing
                columns = ('ID', 'Name', 'Attendance Type', 'Days')
                tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=10)
                
                # Define headings
                tree.heading('ID', text='Employee ID')
                tree.heading('Name', text='Employee Name')
                tree.heading('Attendance Type', text='Attendance Type')
                tree.heading('Days', text='Days')
                
                # Configure column widths to ensure all columns are visible
                total_width = 650  # Approximate table width
                tree.column('ID', width=80, anchor='center', minwidth=60)
                tree.column('Name', width=160, anchor='w', minwidth=120)
                tree.column('Attendance Type', width=280, anchor='center', minwidth=200)
                tree.column('Days', width=80, anchor='center', minwidth=60)
                
                # Set minimum height for table visibility
                tree.configure(height=8)
                
                # Add data to tree
                for emp in employee_data:
                    tree.insert('', tk.END, values=(
                        emp.get('employee_no', 'N/A'),
                        emp.get('employee_name', 'N/A'),
                        emp.get('attendance_type', 'N/A'),
                        emp.get('attendance_days', 'N/A')
                    ))
                
                # Add scrollbar to tree
                tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
                tree.configure(yscrollcommand=tree_scroll.set)
                
                # Pack tree and scrollbar properly
                tree.pack(side="left", fill="both", expand=True)
                tree_scroll.pack(side="right", fill="y")
                
                # Configure table to use full width
                table_frame.columnconfigure(0, weight=1)
                table_frame.rowconfigure(0, weight=1)
            else:
                no_data_label = tk.Label(emp_section, text="❌ No employee data found", 
                        font=('Arial', 12, 'bold'), bg=self.colors['light'], 
                        fg=self.colors['danger'])
                no_data_label.pack(pady=20)
        
        else:  # payroll
            print(f"🎯 DEBUG: Processing payroll display")
            
            # File Information Section
            file_section = tk.LabelFrame(parent_frame, text="� File Information", 
                                       font=('Arial', 11, 'bold'), bg=self.colors['light'],
                                       fg=self.colors['primary'])
            file_section.pack(fill=tk.X, pady=(0, 10), padx=5)
            
            file_info = tk.Frame(file_section, bg=self.colors['light'])
            file_info.pack(fill=tk.X, padx=10, pady=8)
            
            tk.Label(file_info, text=f"📁 File: {result.get('file_name', 'N/A')}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['dark']).pack(anchor=tk.W, pady=2)
            tk.Label(file_info, text=f"📏 Total Rows: {result.get('total_rows', 0)}", 
                    font=('Arial', 10), bg=self.colors['light'], fg=self.colors['dark']).pack(anchor=tk.W, pady=2)
            tk.Label(file_info, text=f"📐 Total Columns: {result.get('total_columns', 0)}", 
                    font=('Arial', 10), bg=self.colors['light'], fg=self.colors['dark']).pack(anchor=tk.W, pady=2)
            
            # Header Information Section
            header_section = tk.LabelFrame(parent_frame, text="📋 Header Information", 
                                         font=('Arial', 11, 'bold'), bg=self.colors['light'],
                                         fg=self.colors['primary'])
            header_section.pack(fill=tk.X, pady=(0, 10), padx=5)
            
            header_info = tk.Frame(header_section, bg=self.colors['light'])
            header_info.pack(fill=tk.X, padx=10, pady=8)
            
            print(f"🎯 DEBUG: Payroll Header - Date: {result.get('date')}, Company: {result.get('company_name')}")
            
            tk.Label(header_info, text=f"� Date: {result.get('date', 'Not found')}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['success']).pack(anchor=tk.W, pady=3)
            tk.Label(header_info, text=f"🏢 Company Name: {result.get('company_name', 'Not found')}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['success']).pack(anchor=tk.W, pady=3)
            tk.Label(header_info, text=f"🏦 Account: {result.get('account', 'Not found')}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['success']).pack(anchor=tk.W, pady=3)
            
            # Payroll Summary Section
            summary_section = tk.LabelFrame(parent_frame, text="💰 Payroll Summary", 
                                          font=('Arial', 11, 'bold'), bg=self.colors['light'],
                                          fg=self.colors['primary'])
            summary_section.pack(fill=tk.X, pady=(0, 10), padx=5)
            
            summary_info = tk.Frame(summary_section, bg=self.colors['light'])
            summary_info.pack(fill=tk.X, padx=10, pady=8)
            
            print(f"🎯 DEBUG: Payroll Summary - Employees: {result.get('total_employees')}, Total: {result.get('total_gross_salary')}")
            
            tk.Label(summary_info, text=f"� Total Employees: {result.get('total_employees', 0)}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['warning']).pack(anchor=tk.W, pady=2)
            tk.Label(summary_info, text=f"💵 Total Gross Salary: {result.get('total_gross_salary', 0):,.2f}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['warning']).pack(anchor=tk.W, pady=2)
            tk.Label(summary_info, text=f"📈 Average Salary: {result.get('average_salary', 0):,.2f}", 
                    font=('Arial', 11, 'bold'), bg=self.colors['light'], fg=self.colors['warning']).pack(anchor=tk.W, pady=2)
            
            # Employee Payroll Data Section
            employee_data = result.get('employee_data', [])
            print(f"🎯 DEBUG: Employee data count: {len(employee_data)}")
            
            emp_section = tk.LabelFrame(parent_frame, text=f"👥 Employee Payroll Data ({len(employee_data)} employees)", 
                                      font=('Arial', 11, 'bold'), bg=self.colors['light'],
                                      fg=self.colors['primary'])
            emp_section.pack(fill=tk.BOTH, expand=True, pady=(0, 10), padx=5)
            
            # Employee payroll data table
            if employee_data:
                print(f"🎯 DEBUG: Creating payroll table with {len(employee_data)} employees")
                
                # Create frame for table with better sizing
                table_frame = tk.Frame(emp_section, bg=self.colors['white'])
                table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
                
                # Create treeview for payroll data - show key salary components
                columns = ('ID', 'Name', 'Basic', 'HRA', 'Medical', 'Total Gross')
                tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=8)
                
                # Define headings
                tree.heading('ID', text='Employee ID')
                tree.heading('Name', text='Employee Name')
                tree.heading('Basic', text='Basic Salary')
                tree.heading('HRA', text='HRA')
                tree.heading('Medical', text='Medical')
                tree.heading('Total Gross', text='Total Gross')
                
                # Configure column widths
                tree.column('ID', width=80, anchor='center', minwidth=60)
                tree.column('Name', width=160, anchor='w', minwidth=120)
                tree.column('Basic', width=100, anchor='e', minwidth=80)
                tree.column('HRA', width=100, anchor='e', minwidth=80)
                tree.column('Medical', width=100, anchor='e', minwidth=80)
                tree.column('Total Gross', width=120, anchor='e', minwidth=100)
                
                # Add data to tree
                for emp in employee_data:
                    salary_components = emp.get('salary_components', {})
                    tree.insert('', tk.END, values=(
                        emp.get('employee_no', 'N/A'),
                        emp.get('employee_name', 'N/A'),
                        f"{salary_components.get('BASIC', 0):,.0f}",
                        f"{salary_components.get('HRA', 0):,.0f}",
                        f"{salary_components.get('MEDICAL', 0):,.0f}",
                        f"{emp.get('total_gross_salary', 0):,.0f}"
                    ))
                
                # Add scrollbar to tree
                tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
                tree.configure(yscrollcommand=tree_scroll.set)
                
                # Pack tree and scrollbar properly
                tree.pack(side="left", fill="both", expand=True)
                tree_scroll.pack(side="right", fill="y")
                
                # Configure table to use full width
                table_frame.columnconfigure(0, weight=1)
                table_frame.rowconfigure(0, weight=1)
            else:
                no_data_label = tk.Label(emp_section, text="❌ No employee payroll data found", 
                        font=('Arial', 12, 'bold'), bg=self.colors['light'], 
                        fg=self.colors['danger'])
                no_data_label.pack(pady=20)
    
    def run(self):
        """Start the application."""
        self.root.mainloop()


if __name__ == "__main__":
    app = ModernExcelProcessor()
    app.run()